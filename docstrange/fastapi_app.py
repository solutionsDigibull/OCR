import asyncio
import json
import os
import shutil
import uuid
from pathlib import Path

import anthropic
from fastapi import BackgroundTasks, FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles

try:
    from analyze_document import (
        POPPLER_PATH,
        extract_page_data,
        merge_results,
        write_csv,
    )
    from pdf2image import convert_from_path
    from PIL import Image
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    DEPS_OK = True
except ImportError as exc:
    print(f"[warn] Missing dependency: {exc}")
    DEPS_OK = False

app = FastAPI(title="DocStrange API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

OUTPUT_DIR = Path("outputs")
OUTPUT_DIR.mkdir(exist_ok=True)

sessions: dict[str, dict] = {}

SUPPORTED = {".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".webp"}


def write_xlsx(data: dict, filepath: str) -> None:
    wb = openpyxl.Workbook()
    h_font = Font(bold=True, color="FFFFFF", size=11)
    h_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    h_align = Alignment(horizontal="center", vertical="center")

    ws1 = wb.active
    ws1.title = "Document Fields"
    ws1.row_dimensions[1].height = 22
    ws1.append(["Field", "Value"])
    for cell in ws1[1]:
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
    for k, v in data.get("key_value_pairs", {}).items():
        ws1.append([str(k), str(v)])
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 55

    for idx, table in enumerate(data.get("tables", []), 1):
        ws = wb.create_sheet(title=table.get("title", f"Table {idx}")[:31])
        headers = table.get("headers", [])
        rows = table.get("rows", [])
        if headers:
            ws.append(headers)
            for cell in ws[1]:
                cell.font = h_font
                cell.fill = h_fill
                cell.alignment = h_align
        for row in rows:
            ws.append([str(c) for c in row])
        for col in ws.columns:
            width = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(width + 4, 60)

    ws_notes = wb.create_sheet("Notes")
    ws_notes.append(["#", "Note"])
    for cell in ws_notes[1]:
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
    for j, note in enumerate(data.get("notes", []), 1):
        ws_notes.append([j, str(note)])
    ws_notes.column_dimensions["A"].width = 6
    ws_notes.column_dimensions["B"].width = 90

    wb.save(filepath)


def _sync_extraction(session_id: str, input_path: str, filename: str) -> None:
    try:
        client = anthropic.Anthropic()
        suffix = Path(filename).suffix.lower()

        if suffix == ".pdf":
            poppler = POPPLER_PATH if os.path.exists(POPPLER_PATH) else None
            pages = convert_from_path(input_path, dpi=200, poppler_path=poppler)
        elif suffix in {".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".webp"}:
            pages = [Image.open(input_path)]
        else:
            raise ValueError(f"Unsupported format: {suffix}")

        total = len(pages)
        page_results = []
        for i, page_img in enumerate(pages, 1):
            sessions[session_id]["progress"] = int((i - 1) / total * 80)
            sessions[session_id]["status_message"] = f"Analyzing page {i} of {total}…"
            result = extract_page_data(client, page_img, i)
            page_results.append(result)

        sessions[session_id]["status_message"] = "Generating output files…"
        sessions[session_id]["progress"] = 90

        data = merge_results(page_results)
        base = session_id
        csv_path = str(OUTPUT_DIR / f"{base}.csv")
        xlsx_path = str(OUTPUT_DIR / f"{base}.xlsx")
        json_path = str(OUTPUT_DIR / f"{base}.json")

        write_csv(data, csv_path)
        write_xlsx(data, xlsx_path)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        sessions[session_id].update(
            {
                "status": "completed",
                "progress": 100,
                "status_message": "Extraction complete",
                "data": data,
                "csv_path": csv_path,
                "xlsx_path": xlsx_path,
                "json_path": json_path,
                "stats": {
                    "field_count": len(data.get("key_value_pairs", {})),
                    "table_count": len(data.get("tables", [])),
                    "note_count": len(data.get("notes", [])),
                },
            }
        )
    except Exception as exc:
        sessions[session_id].update(
            {"status": "failed", "status_message": str(exc), "error": str(exc)}
        )
    finally:
        if os.path.exists(input_path):
            os.unlink(input_path)


async def run_extraction(session_id: str, input_path: str, filename: str) -> None:
    await asyncio.to_thread(_sync_extraction, session_id, input_path, filename)


@app.post("/upload")
async def upload_file(
    background_tasks: BackgroundTasks, file: UploadFile = File(...)
):
    suffix = Path(file.filename).suffix.lower()
    if suffix not in SUPPORTED:
        raise HTTPException(
            400,
            f"Unsupported format '{suffix}'. Supported: {', '.join(sorted(SUPPORTED))}",
        )

    session_id = str(uuid.uuid4())
    input_path = str(OUTPUT_DIR / f"{session_id}_input{suffix}")

    with open(input_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    file_size = os.path.getsize(input_path)

    sessions[session_id] = {
        "id": session_id,
        "status": "processing",
        "progress": 0,
        "status_message": "Starting extraction…",
        "filename": file.filename,
        "file_size": file_size,
        "data": None,
        "error": None,
        "stats": None,
    }

    background_tasks.add_task(run_extraction, session_id, input_path, file.filename)

    return {
        "id": session_id,
        "status": "processing",
        "filename": file.filename,
        "file_size": file_size,
    }


@app.get("/results/{id}")
async def get_results(id: str):
    session = sessions.get(id)
    if not session:
        raise HTTPException(404, f"Session '{id}' not found")
    return {
        "id": id,
        "status": session["status"],
        "progress": session.get("progress", 0),
        "status_message": session.get("status_message", ""),
        "filename": session.get("filename"),
        "file_size": session.get("file_size"),
        "data": session.get("data"),
        "stats": session.get("stats"),
        "error": session.get("error"),
    }


@app.get("/download/{id}")
async def download_file(id: str, format: str = "csv"):
    session = sessions.get(id)
    if not session:
        raise HTTPException(404, "Session not found")
    if session["status"] != "completed":
        raise HTTPException(400, "Results not ready yet")

    fmt = format.lower()
    path_key = f"{fmt}_path"
    path = session.get(path_key)

    media_types = {
        "csv": "text/csv",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "json": "application/json",
    }
    if fmt not in media_types:
        raise HTTPException(400, f"Unknown format '{fmt}'. Use csv, xlsx, or json.")
    if not path or not os.path.exists(path):
        raise HTTPException(404, "Output file not found on disk")

    stem = Path(session["filename"]).stem
    return FileResponse(
        path,
        media_type=media_types[fmt],
        filename=f"{stem}_extracted.{fmt}",
    )


@app.get("/health")
async def health():
    return {"status": "ok", "version": "1.0.0"}


# Serve built React app in production
_dist = Path(__file__).parent / "frontend" / "dist"
if _dist.exists():
    app.mount("/assets", StaticFiles(directory=str(_dist / "assets")), name="assets")

    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        idx = _dist / "index.html"
        if idx.exists():
            return FileResponse(str(idx))
        raise HTTPException(404, "Frontend not built. Run: cd frontend && npm run build")


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("fastapi_app:app", host="0.0.0.0", port=8000, reload=True)
