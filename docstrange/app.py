from fastapi import FastAPI, Query, UploadFile, File
from pydantic import BaseModel
from typing import List
from fastapi.responses import FileResponse, JSONResponse

import os
import cv2
import numpy as np
from pdf2image import convert_from_path
import pytesseract
from openpyxl import Workbook
from datetime import datetime
import re
import shutil

# -------------------------------
# CONFIG
# -------------------------------
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
POPPLER_PATH = r"C:\poppler\Library\bin"

OUTPUT_DIR = "outputs"
UPLOAD_DIR = "uploads"

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = FastAPI()

# -------------------------------
# REQUEST MODEL
# -------------------------------
class BatchFileRequest(BaseModel):
    file_paths: List[str]

# -------------------------------
# LOAD FILES (PDF + IMAGE)
# -------------------------------
def load_images(file_path):
    ext = file_path.lower().split(".")[-1]

    if ext == "pdf":
        return convert_from_path(file_path, dpi=300, poppler_path=POPPLER_PATH)

    elif ext in ["png", "jpg", "jpeg", "bmp", "tiff", "webp"]:
        img = cv2.imread(file_path)
        if img is None:
            raise Exception(f"Cannot read image: {file_path}")
        return [img]

    else:
        raise Exception(f"Unsupported file type: {file_path}")

# -------------------------------
# TABLE DETECTION
# -------------------------------
def detect_tables(image):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    thresh = cv2.adaptiveThreshold(
        gray, 255,
        cv2.ADAPTIVE_THRESH_MEAN_C,
        cv2.THRESH_BINARY_INV,
        15, 4
    )

    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    tables = []
    for cnt in contours:
        x, y, w, h = cv2.boundingRect(cnt)

        if w > 300 and h > 150:
            tables.append(image[y:y+h, x:x+w])

    return tables

# -------------------------------
# VALIDATORS
# -------------------------------
def is_valid_part_no(text):
    return bool(re.match(r'^[A-Za-z0-9\-]+$', text))

def is_valid_qty(text):
    return text.isdigit()

def is_valid_matl(text):
    return bool(re.match(r'^[A-Za-z.]{2,}$', text))

# -------------------------------
# STRUCTURED TEXT
# -------------------------------
def extract_structured_lines(image):
    text = pytesseract.image_to_string(image)

    lines = []
    for line in text.split("\n"):
        line = line.strip()
        if re.match(r'.+\(\d+\)$', line):
            lines.append(line)

    return lines

# -------------------------------
# TABLE EXTRACTION
# -------------------------------
def extract_table(region):
    data = pytesseract.image_to_data(region, output_type=pytesseract.Output.DICT)

    rows = []
    current_y = -100

    for i in range(len(data["text"])):
        word = data["text"][i].strip()
        if not word:
            continue

        x = data["left"][i]
        y = data["top"][i]

        if abs(y - current_y) > 12:
            rows.append([])
            current_y = y

        rows[-1].append((x, word))

    rows = [sorted(r, key=lambda x: x[0]) for r in rows]

    # detect header (top or bottom)
    header_index = -1
    for i, row in enumerate(rows):
        txt = " ".join([w for _, w in row]).lower()

        if (
            "name" in txt and
            ("no" in txt or "part" in txt) and
            ("mat" in txt or "material" in txt) and
            ("q" in txt or "qty" in txt)
        ):
            header_index = i
            break

    if header_index != -1:
        if header_index > len(rows) // 2:
            rows = rows[:header_index]
            rows.reverse()
        else:
            rows = rows[header_index + 1:]

    structured = []

    COL_PART_NO = 150
    COL_NAME = 400
    COL_MATL = 650

    for row in rows:
        part_no, part_name, matl, qty = [], [], [], []

        for x, word in row:
            if x < COL_PART_NO:
                part_no.append(word)
            elif x < COL_NAME:
                part_name.append(word)
            elif x < COL_MATL:
                matl.append(word)
            else:
                qty.append(word)

        part_no = " ".join(part_no).strip()
        part_name = " ".join(part_name).strip()
        matl = " ".join(matl).strip()
        qty = " ".join(qty).strip()

        row_text = f"{part_no} {part_name} {matl} {qty}".lower()

        if not row_text:
            continue

        if all(c in "+-.$0123456789 " for c in row_text):
            continue

        if not is_valid_part_no(part_no):
            part_no = ""

        if not part_name or len(part_name) < 2:
            continue

        if not is_valid_matl(matl):
            matl = ""

        if not is_valid_qty(qty):
            qty = ""

        structured.append([part_no, part_name, matl, qty])

    return structured

# -------------------------------
# PROCESS FILE
# -------------------------------
def process_file(file_path):
    images = load_images(file_path)

    tables_found = []
    structured_text_found = []

    for img in images:
        if not isinstance(img, np.ndarray):
            img = np.array(img)

        regions = detect_tables(img)

        for region in regions:
            table = extract_table(region)
            if table:
                tables_found.append(table)

        structured_text_found.extend(extract_structured_lines(img))

    name = os.path.basename(file_path)

    return name, {
        "tables": tables_found,
        "structured_text": structured_text_found
    }

# -------------------------------
# EXCEL
# -------------------------------
def create_excel(all_data):
    wb = Workbook()

    for file_name, content in all_data.items():
        ws = wb.create_sheet(file_name[:31])

        if content["tables"]:
            ws.append(["Part No", "Part Name", "Material", "Qty"])
            for table in content["tables"]:
                for row in table:
                    ws.append(row)
                ws.append([])

        if content["structured_text"]:
            ws.append([])
            ws.append(["Text Extraction"])
            for line in content["structured_text"]:
                ws.append([line])

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    path = os.path.join(OUTPUT_DIR, f"output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(path)
    return path

# -------------------------------
# ✅ BATCH (Swagger safe)
# -------------------------------
@app.post("/process-batch", response_class=JSONResponse)
def process_batch(
    input: BatchFileRequest,
    format: str = Query("json", enum=["json", "xlsx"])
):
    all_data = {}

    for file_path in input.file_paths:
        name, result = process_file(file_path)
        all_data[name] = result

    if format == "xlsx":
        excel_path = create_excel(all_data)
        return FileResponse(
            excel_path,
            filename=os.path.basename(excel_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    return {"status": "success", "data": all_data}

# -------------------------------
# ✅ UPLOAD API (FOR ACCOMPLISH)
# -------------------------------
@app.post("/process-upload")
async def process_upload(files: List[UploadFile] = File(...)):
    all_data = {}

    for file in files:
        path = os.path.join(UPLOAD_DIR, file.filename)

        with open(path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        name, result = process_file(path)
        all_data[name] = result

    return {
        "status": "success",
        "data": all_data
    }

# -------------------------------
# HEALTH
# -------------------------------
@app.get("/")
def home():
    return {"status": "running 🚀"}