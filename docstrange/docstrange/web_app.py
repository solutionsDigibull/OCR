"""Web application for docstrange document extraction."""

import csv as csv_module
import io
import os
import json
import re
import tempfile
import threading
import uuid
from pathlib import Path
from typing import Optional
from flask import Flask, request, jsonify, render_template, send_from_directory
from werkzeug.utils import secure_filename
from werkzeug.exceptions import RequestEntityTooLarge

from .extractor import DocumentExtractor
from .exceptions import ConversionError, UnsupportedFormatError, FileNotFoundError

# In-memory session store for async extraction
_sessions: dict = {}
_sessions_lock = threading.Lock()


def _parse_markdown_to_structured(markdown_text: str):
    """Parse markdown output into key_value_pairs, tables, and notes."""
    key_value_pairs = {}
    tables = []
    notes = []

    lines = markdown_text.split('\n')
    i = 0
    current_section = ''

    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        # Section heading
        if stripped.startswith('#'):
            current_section = stripped.lstrip('#').strip()
            i += 1
            continue

        # Markdown table (next line is separator)
        if stripped.startswith('|') and i + 1 < len(lines) and re.match(r'^\s*\|[\s\-:|]+\|\s*$', lines[i + 1]):
            table_lines = []
            while i < len(lines) and lines[i].strip().startswith('|'):
                table_lines.append(lines[i].strip())
                i += 1
            if len(table_lines) >= 3:
                def _split_row(r):
                    parts = r.split('|')
                    return [c.strip() for c in parts[1:-1]] if len(parts) > 2 else [c.strip() for c in parts if c.strip()]
                headers = _split_row(table_lines[0])
                rows = [_split_row(r) for r in table_lines[2:] if _split_row(r)]
                kv_headers = {'field', 'key', 'label', 'parameter', 'name', 'item', 'property'}
                val_headers = {'value', 'data', 'result', 'description', 'desc'}
                is_kv = (
                    len(headers) == 2 and (
                        headers[0].lower() in kv_headers or headers[1].lower() in val_headers
                    )
                )
                if is_kv:
                    for row in rows:
                        if len(row) >= 2:
                            key_value_pairs[row[0]] = row[1]
                        elif len(row) == 1 and row[0]:
                            key_value_pairs[row[0]] = ''
                else:
                    tables.append({
                        'title': current_section or f'Table {len(tables) + 1}',
                        'headers': headers,
                        'rows': rows,
                    })
            continue

        # **key**: value  or  **key** - value
        m = re.match(r'\*\*(.+?)\*\*\s*[:\-]\s*(.*)', stripped)
        if m:
            key_value_pairs[m.group(1).strip()] = m.group(2).strip()
            i += 1
            continue

        # key: value  (simple colon pair, not a heading, not a URL)
        m = re.match(r'^([A-Za-z][A-Za-z0-9 /\-_()\[\].]{1,50}):\s+([^\n]{1,200})$', stripped)
        if m and not stripped.startswith('#') and '://' not in stripped:
            key_value_pairs[m.group(1).strip()] = m.group(2).strip()
            i += 1
            continue

        # Bullet / numbered list → notes
        m = re.match(r'^[-*•]\s+(.+)$', stripped)
        if m:
            notes.append(m.group(1))
            i += 1
            continue
        m = re.match(r'^\d+[.)]\s+(.+)$', stripped)
        if m:
            notes.append(m.group(1))
            i += 1
            continue

        i += 1

    stats = {
        'field_count': len(key_value_pairs),
        'table_count': len(tables),
        'note_count': len(notes),
    }
    return {'key_value_pairs': key_value_pairs, 'tables': tables, 'notes': notes}, stats


def _run_extraction(session_id: str, tmp_path: str, filename: str, file_size: int):
    """Background worker: extract document and store result in session."""
    def _update(**kwargs):
        with _sessions_lock:
            _sessions[session_id].update(kwargs)

    try:
        _update(progress=10, status_message='Starting extraction…')
        extractor = DocumentExtractor()
        _update(progress=30, status_message='Running OCR…')
        result = extractor.extract(tmp_path)
        _update(progress=70, status_message='Parsing structure…')
        markdown = result.extract_markdown()
        data, stats = _parse_markdown_to_structured(markdown)
        _update(status='completed', progress=100, status_message='Done!',
                data=data, stats=stats, markdown=markdown)
    except Exception as exc:
        _update(status='failed', error=str(exc))
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size

def check_gpu_availability():
    """Check if GPU is available for processing."""
    try:
        import torch
        return torch.cuda.is_available()
    except ImportError:
        return False

def download_models():
    """Download models synchronously before starting the app."""
    print("Starting model download...")

    gpu_available = check_gpu_availability()

    if gpu_available:
        print("GPU detected - downloading GPU models")
        extractor = DocumentExtractor(gpu=True)
    else:
        print("GPU not available - using cloud processing")
        extractor = DocumentExtractor()

    print("Downloading models...")

    test_content = "Test document for model download."
    with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False) as tmp_file:
        tmp_file.write(test_content)
        test_file_path = tmp_file.name

    try:
        result = extractor.extract(test_file_path)
        print("Model download completed successfully")
    except Exception as e:
        print(f"Model download warning: {e}")
    finally:
        if os.path.exists(test_file_path):
            os.unlink(test_file_path)

def create_extractor_with_mode(processing_mode):
    """Create DocumentExtractor with proper error handling for processing mode."""
    if processing_mode == 'gpu':
        if not check_gpu_availability():
            raise ValueError("GPU mode selected but GPU is not available. Please install PyTorch with CUDA support.")
        return DocumentExtractor(gpu=True)
    else:  # cloud mode (default)
        return DocumentExtractor()

# Initialize the document extractor
extractor = DocumentExtractor()

@app.route('/health')
def health():
    return jsonify({'status': 'healthy'})


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'detail': 'No file provided'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'detail': 'No file selected'}), 400

    session_id = str(uuid.uuid4())
    filename = secure_filename(file.filename)
    suffix = Path(filename).suffix

    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        file.save(tmp.name)
        tmp_path = tmp.name

    file_size = os.path.getsize(tmp_path)

    with _sessions_lock:
        _sessions[session_id] = {
            'status': 'processing',
            'progress': 0,
            'status_message': 'Starting…',
            'filename': filename,
            'file_size': file_size,
        }

    t = threading.Thread(target=_run_extraction, args=(session_id, tmp_path, filename, file_size), daemon=True)
    t.start()

    return jsonify({'id': session_id, 'filename': filename})


@app.route('/results/<session_id>')
def get_session_results(session_id):
    with _sessions_lock:
        session = dict(_sessions.get(session_id, {}))
    if not session:
        return jsonify({'detail': 'Session not found'}), 404
    # Don't send raw markdown over the wire
    session.pop('markdown', None)
    return jsonify(session)


@app.route('/download/<session_id>')
def download_results(session_id):
    with _sessions_lock:
        session = dict(_sessions.get(session_id, {}))
    if not session or session.get('status') != 'completed':
        return jsonify({'detail': 'Results not ready'}), 404

    fmt = request.args.get('format', 'json')
    data = session.get('data', {})
    stem = Path(session.get('filename', 'document')).stem

    if fmt == 'json':
        content = json.dumps(data, indent=2)
        return app.response_class(
            response=content,
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment; filename="{stem}.json"'},
        )

    if fmt == 'csv':
        buf = io.StringIO()
        writer = csv_module.writer(buf)
        writer.writerow(['Field', 'Value'])
        for k, v in data.get('key_value_pairs', {}).items():
            writer.writerow([k, v])
        return app.response_class(
            response=buf.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename="{stem}.csv"'},
        )

    if fmt == 'xlsx':
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Fields'
            ws.append(['Field', 'Value'])
            for k, v in data.get('key_value_pairs', {}).items():
                ws.append([k, v])
            for idx, tbl in enumerate(data.get('tables', [])):
                ws2 = wb.create_sheet(title=(tbl.get('title') or f'Table {idx+1}')[:31])
                if tbl.get('headers'):
                    ws2.append(tbl['headers'])
                for row in tbl.get('rows', []):
                    ws2.append(row)
            if data.get('notes'):
                ws3 = wb.create_sheet(title='Notes')
                ws3.append(['#', 'Note'])
                for idx, note in enumerate(data['notes'], 1):
                    ws3.append([idx, note])
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return app.response_class(
                response=buf.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename="{stem}.xlsx"'},
            )
        except ImportError:
            return jsonify({'detail': 'openpyxl not installed'}), 500

    return jsonify({'detail': 'Unknown format'}), 400


@app.route('/')
def index():
    """Serve the main page."""
    return render_template('index.html')

@app.route('/static/<path:filename>')
def static_files(filename):
    """Serve static files."""
    return send_from_directory('static', filename)

@app.route('/api/extract', methods=['POST'])
def extract_document():
    """API endpoint for document extraction."""
    try:
        # Check if file was uploaded
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Get parameters
        output_format = request.form.get('output_format', 'markdown')
        processing_mode = request.form.get('processing_mode', 'cloud')
        
        # Create extractor based on processing mode
        try:
            extractor = create_extractor_with_mode(processing_mode)
        except ValueError as e:
            return jsonify({'error': str(e)}), 400
        
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix=Path(file.filename).suffix) as tmp_file:
            file.save(tmp_file.name)
            tmp_path = tmp_file.name
        
        try:
            # Extract content
            result = extractor.extract(tmp_path)
            
            # Convert to requested format
            if output_format == 'markdown':
                content = result.extract_markdown()
            elif output_format == 'html':
                content = result.extract_html()
            elif output_format == 'json':
                content = result.extract_data()
                content = json.dumps(content, indent=2)
            elif output_format == 'csv':
                content = result.extract_csv(include_all_tables=True)
            elif output_format == 'flat-json':
                content = result.extract_data()
                content = json.dumps(content, indent=2)
            elif output_format == 'text':
                content = result.extract_text()
            else:
                content = result.extract_markdown()  # Default to markdown
            
            # Get metadata
            metadata = {
                'file_type': Path(file.filename).suffix.lower(),
                'file_name': file.filename,
                'file_size': os.path.getsize(tmp_path),
                'pages_processed': getattr(result, 'pages_processed', 1),
                'processing_time': getattr(result, 'processing_time', 0),
                'output_format': output_format,
                'processing_mode': processing_mode
            }
            
            return jsonify({
                'success': True,
                'content': content,
                'metadata': metadata
            })
            
        finally:
            # Clean up temporary file
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
                
    except RequestEntityTooLarge:
        return jsonify({'error': 'File too large. Maximum size is 100MB.'}), 413
    except UnsupportedFormatError as e:
        return jsonify({'error': f'Unsupported file format: {str(e)}'}), 400
    except ConversionError as e:
        return jsonify({'error': f'Conversion error: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': f'Unexpected error: {str(e)}'}), 500

@app.route('/api/supported-formats')
def get_supported_formats():
    """Get list of supported file formats."""
    formats = extractor.get_supported_formats()
    return jsonify({'formats': formats})

@app.route('/api/health')
def health_check():
    """Health check endpoint."""
    return jsonify({'status': 'healthy', 'version': '1.0.0'})

@app.route('/api/system-info')
def get_system_info():
    """Get system information including GPU availability."""
    gpu_available = check_gpu_availability()
    
    # Get additional system info
    system_info = {
        'gpu_available': gpu_available,
        'processing_modes': {
            'cloud': {
                'available': True,
                'description': 'Process using cloud API. Fast and requires no local setup.'
            },
            'gpu': {
                'available': gpu_available,
                'description': 'Process locally using GPU. Fastest local processing, requires CUDA.' if gpu_available else 'GPU not available. Install PyTorch with CUDA support.'
            }
        }
    }
    
    return jsonify(system_info)

def run_web_app(host='0.0.0.0', port=8000, debug=False):
    """Run the web application."""
    gpu_available = check_gpu_availability()
    if gpu_available:
        print("GPU detected - downloading models before starting...")
        download_models()
    else:
        print("No GPU detected - starting in cloud processing mode.")
    print(f"Starting docstrange web interface at http://{host}:{port}")
    print("Press Ctrl+C to stop the server")
    app.run(host=host, port=port, debug=debug)

if __name__ == '__main__':
    run_web_app(debug=True) 