import csv
import os
from openpyxl import Workbook

CSV_PATH = "drawing_data.csv"
OUTPUT_DIR = "outputs"

os.makedirs(OUTPUT_DIR, exist_ok=True)

rows = []
with open(CSV_PATH, newline="", encoding="utf-8") as f:
    for r in csv.DictReader(f):
        if r["Category"]:
            rows.append(r)

title_block = {r["Field"]: r["Value"] for r in rows if r["Category"] == "Title Block"}
dims = [f'{r["Field"]} – {r["Value"]}' for r in rows if r["Category"] == "Dimensions"]
annotations = [f'{r["Field"]}: {r["Value"]}' for r in rows if r["Category"] == "Handwritten Annotations"]

parts = [{
    "number": title_block.get("Drawing Number", ""),
    "name":   title_block.get("Part Type", "")
}]

# drawing_data_output.xlsx — 3 sheets matching create_excel() in app.py
wb = Workbook()
ws1 = wb.active
ws1.title = "Parts"
ws1["A1"] = "Part No"
ws1["B1"] = "Part Name"
for i, p in enumerate(parts, 2):
    ws1[f"A{i}"] = p["number"]
    ws1[f"B{i}"] = p["name"]

ws2 = wb.create_sheet("Dimensions")
ws2["A1"] = "Values"
for i, d in enumerate(dims, 2):
    ws2[f"A{i}"] = d

ws3 = wb.create_sheet("Annotations")
ws3["A1"] = "Text"
for i, a in enumerate(annotations, 2):
    ws3[f"A{i}"] = a

output_path = os.path.join(OUTPUT_DIR, "drawing_data_output.xlsx")
wb.save(output_path)
print(f"Saved: {output_path}")

# drawing_data_parts.xlsx — full structured table (all CSV rows)
wb2 = Workbook()
ws = wb2.active
ws.title = "Parts List"
ws.append(["Category", "Field", "Value"])
for r in rows:
    ws.append([r["Category"], r["Field"], r["Value"]])

parts_path = os.path.join(OUTPUT_DIR, "drawing_data_parts.xlsx")
wb2.save(parts_path)
print(f"Saved: {parts_path}")
